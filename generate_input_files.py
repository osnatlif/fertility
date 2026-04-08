"""Generate new input files from Excel for 1960white and 1985white cohorts."""
import openpyxl

excel_path = r'C:\Users\Osnat\Dropbox\Academic Research\Rania Gihleb\Rania\Education, Marriage and Fertility\data\out files STATA\models moments 60 and 85\1960 1985 moments + initial values.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def sheet_to_list(ws, skip=1):
    rows = []
    for row in ws.iter_rows(min_row=skip+1, values_only=True):
        vals = list(row)
        if any(v is not None for v in vals):
            rows.append(vals)
    return rows

# MARRIAGE/DIVORCE
ws = wb['marriage']
rows = sheet_to_list(ws)
for cy in [1960, 1985]:
    cn = str(cy) + 'white'
    for gender, label in [('w', 'female'), ('h', 'male')]:
        with open(f'input/marr_divorce_{gender}{cn}.txt', 'w') as f:
            for r in rows:
                if r[0] == cy and r[1] == label:
                    f.write(f'{r[2]}\t{r[3]}\t{r[4]:.7f}\t{r[5]:.7f}\n')
print('Marriage/divorce files created')

# EMPLOYMENT/WAGE
ws = wb['emp wage']
rows = sheet_to_list(ws)
for cy in [1960, 1985]:
    cn = str(cy) + 'white'
    for gender, label in [('w', 'female'), ('h', 'male')]:
        for marital, mlabel in [('married', 'married'), ('unmarried', 'unmarried')]:
            fname = f'input/{marital}_{gender}{cn}.txt'
            with open(fname, 'w') as f:
                for r in rows:
                    match_marital = r[2] == mlabel or (mlabel == 'unmarried' and r[2] == 0)
                    if r[0] == cy and r[1] == label and match_marital:
                        educ = r[3]
                        ag = r[4]
                        ft = r[6] if r[6] is not None else 0
                        pt = r[7] if r[7] is not None else 0
                        wage = r[8] if r[8] is not None else 0
                        welfare = r[9] if r[9] is not None else 0
                        f.write(f'{educ}\t{ag}\t{ft:.7f}\t{pt:.7f}\t{wage:.2f}\t{welfare:.7f}\n')
print('Employment/wage files created')

# FERTILITY
ws = wb['fertility']
all_rows = []
for row in ws.iter_rows(min_row=3, values_only=True):
    all_rows.append(list(row))

for cy in [1960, 1985]:
    cn = str(cy) + 'white'
    with open(f'input/fertility_unmarried{cn}.txt', 'w') as f:
        for r in all_rows:
            if r[1] == cy:
                f.write(f'{r[2]}\t{r[3]}\t{r[4]:.6f}\t{r[5]:.7f}\n')
    with open(f'input/fertility_married{cn}.txt', 'w') as f:
        for r in all_rows:
            if r[7] == cy:
                f.write(f'{r[8]}\t{r[9]}\t{r[10]:.6f}\t{r[11]:.7f}\n')
print('Fertility files created')

# ASSORTATIVE MATING
ws = wb['assortative']
all_rows = []
for row in ws.iter_rows(min_row=1, values_only=True):
    all_rows.append(list(row))

men_start = women_start = None
for i, r in enumerate(all_rows):
    if r[0] == 'Men': men_start = i
    if r[0] == 'Women': women_start = i

def write_assortative(f, rows_section, cy):
    for r in rows_section:
        if r[0] == cy:
            ag = r[1]
            # CG: cols 2,3,4 (wife/husband hs,sc,cg); SC: cols 8,9,10; HS: cols 14,15,16
            for edu_cols in [(14,15,16), (8,9,10), (2,3,4)]:  # HS, SC, CG order
                vals = [r[c] if r[c] is not None else 0 for c in edu_cols]
                total = sum(vals)
                if total > 0:
                    cond = [v/total for v in vals]
                    f.write(f'{ag}\t{cond[0]:.7f}\t{cond[1]:.7f}\t{cond[2]:.7f}\n')
                else:
                    f.write(f'{ag}\t0\t0\t0\n')

for cy in [1960, 1985]:
    cn = str(cy) + 'white'
    with open(f'input/assortative_h{cn}.txt', 'w') as f:
        write_assortative(f, all_rows[men_start+2:women_start], cy)
    with open(f'input/assortative_w{cn}.txt', 'w') as f:
        write_assortative(f, all_rows[women_start+2:], cy)
print('Assortative mating files created')

# EDUCATION DISTRIBUTION
ws = wb['education']
rows = sheet_to_list(ws)
for cy in [1960, 1985]:
    cn = str(cy) + 'white'
    for gender, label in [('w', 'female'), ('h', 'male')]:
        with open(f'input/education_{gender}{cn}.txt', 'w') as f:
            for r in rows:
                if r[0] == cy and r[3] == label and r[1] == 1:
                    f.write(f'{r[2]}\t{r[4]:.7f}\t{r[5]:.7f}\t{r[6]:.7f}\n')
print('Education distribution files created')

print('\nAll input files generated successfully!')
